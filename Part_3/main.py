#%matplotlib inline
import pandas as pd
import os
from openpyxl import load_workbook
from datetime import datetime
from sklearn.cross_validation import train_test_split
import numpy as np


# Chemin vers le dossier ou sont les codes du projet
path_code = '/mnt/smb/TAMPON/Igor/RFR/script/'
os.chdir(path_code)
import data_preparation
import K_means as K_means
import IntConf
import prediction

########################
##Debut de parametrage##
########################
## Documentation du main 
#  @brief Paramétrage

print("ooooooooooooooooooooooooooo")
# Chemin vers le dossier ou sont les données


path="/mnt/smb/TAMPON/Igor/RFR/"
path_data = "/mnt/smb/TAMPON/Partages/RFR/matrice de donnees/"
#path_data = path + "data_src/"
path_rslt = path + "data_rslt/"
os.chdir(path_data)

R_cont_y = 0.25; R_Cramer_y = 0.20; R_cont_x = 0.85; R_Cramer_x = 0.80
taille_ech = 33363; nb_clusters = 33
method_disc = 'Cramer'; method_continuous = 'regression'
method_prediction = 'random_forest'
fichier = 'variables_explicatives_BQ_SD_201703.csv' 
fichier_sortie = 'revenu_client_BQ_SD_201703.csv'
dic={'SGMT_PF_V4':['SGMT_PF_V4',1],'SGMT_PF_AXE_FIDELITE_V4':['SGMT_PF_AXE_FIDELITE_V4',3],'REVENU_EST_M':['REVENU_EST_M',3]}
#dic={'A':['A',1],'race':['race',1]}



########################
##Fin de parametrage##
########################

#################################
###### Début de compilation #####
#################################

begin = datetime.now()
begin_global = datetime.now()

#============================================================================
#                 Préparation des données
dfX, Y, R_dico = data_preparation.main(
            fichier = fichier,fichier_sortie= fichier_sortie,method_disc=method_disc,method_continuous=method_continuous,taille_ech = taille_ech,
            R_cont_y = R_cont_y , R_Cramer_y = R_Cramer_y, R_cont_x = R_cont_x, R_Cramer_x = R_Cramer_x, dic=dic,
            normalize = True, verbose = True, path_rslt)

delta_time = round((datetime.now() - begin).total_seconds(),1)
print('Données préparées en ' + str(delta_time) + 's')

#============================================================================
#               K-means
begin = datetime.now()
repartition = K_means.K_means(dfX,Y,nb_clusters_init= nb_clusters,methode_prediction=method_prediction)
delta_time = round((datetime.now() - begin).total_seconds(),1)
print('Clusters effectués en ' + str(delta_time) + 's')

#===========================================================================
#              IC
begin = datetime.now()
cluster_stat = IntConf.IntConf(repartition,alpha = 0.85,path_rslt)
print(cluster_stat)
delta_time = round((datetime.now() - begin).total_seconds(),1)
print('Intervalle de confiance éffectué en ' + str(delta_time) + 's')

#===========================================================================
#              Prediction
index = ~Y.isnull()
dfX_train, dfX_test, Y_train, Y_test = train_test_split(dfX[index], Y[index],test_size = 0.4, random_state = 44)

del dfX_train, dfX_test, Y_train
result, result_test=prediction.IC_reg(repartition, dfX, Y,path_rslt)
result_1 = prediction.get_result(repartition, dfX, Y, Y_test, method = method_prediction)
#prediction.regression_graph(Y_real = result_1['Y_real'], Y_pred = result_1['Y_pred'], col = 'black')
score_abs, score_rel = prediction.get_regression_score(Y_real = result_1['Y_real'], Y_pred = result_1['Y_pred'])
print('Erreur moyenne de prevision: ' + str(score_abs) + ' (' + str(score_rel) + '%)')
curve_Recall, AUC_ROC = prediction.classification_curve(Y_real = result_1['Y_real'], Y_pred = result_1['Y_pred'],
                   threshold_rich = 100000, col = 'black')
print('AUC_ROC: '+str(AUC_ROC))
print('AUC Precision-Recall: '+str(curve_Recall))
#prediction.graph_variable_influence(dfX,result_1,col='black')

#===========================================================================
#              file of parameters 

if os.path.exists(path+"Compare.xlsx") == False:
    S=pd.DataFrame(columns=["reference","data_x","data_y","method disc","method continuous","method prevision","Date",
    "Nb ligne","% NaN","k cluster","R_cont_y","R_Cramer_y","R_cont_x","R_Cramer_x","Nb var quali","Nb var quant",
    "Nb regroupement","score","score relative","AUC ROC","curve Recall","temps de calcul"])
    S.to_excel(path+"Compare.xlsx",encoding="utf-8", index=False) #,sep=";",encoding="utf-8"

Nb_var_quant=len(R_dico['variables continues gardees']);
Nb_var_quali=len(R_dico['variables discretes gardees']);
Nb_regroupement=len(R_dico['groupe variables'])
date = "%s" % datetime.now()
NaN = len(Y[Y.isnull()==True])/Y.shape[0]
delta_time = round((datetime.now() - begin_global).total_seconds(),1)

wb = load_workbook(path+"Compare.xlsx")
sheet = wb.get_sheet_by_name('Sheet1')
reference = sheet.max_row + 1
ABC=list("ABCDEFGHIJKLMNOPQRST")
liste=[reference,fichier,fichier_sortie,method_disc,method_continuous,method_prediction,date,taille_ech,str(NaN.__round__(2)),
       nb_clusters,R_cont_y,R_Cramer_y,R_cont_x,R_Cramer_x,Nb_var_quali,Nb_var_quant,Nb_regroupement,score_abs,
       str(score_rel)+'%',str(AUC_ROC),str(curve_Recall),str(delta_time) + 's']
for i in range(len(ABC)):
    sheet[ABC[i] + str(reference)].value = liste[i]
	
wb.save(path+"Compare.xlsx")

mon_fichier = open(path_rslt + "fichier.txt", "w")
mon_fichier.write('liste variables qualitatives :')
a=R_dico['variables discretes gardees'].sort(['R2'],ascending=[False])
for i in a.values:
    mon_fichier.write('\n        '+str(i))

a=R_dico['variables continues gardees'].sort(['R2'],ascending=[False])    
mon_fichier.write('\n \nliste variables quantitatives :')
for i in a.values:
    mon_fichier.write('\n        '+str(i))
   
mon_fichier.write('\n \nliste groupe variables :')
mon_fichier.write('\n          R²,    groupe variables correlees,    variable representante')
for i in R_dico['groupe variables'].values:
    mon_fichier.write('\n        '+str(i))

delta_time = round((datetime.now() - begin_global).total_seconds(),1)    
mon_fichier.write('\n \nTemps de calcul est ' + str(delta_time)+ 's')
mon_fichier.close() 